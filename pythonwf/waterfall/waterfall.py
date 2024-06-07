from typing import Dict, List, Any, Optional
from pythonwf.construct_sql.construct_sql import SQLConstructor
from pythonwf.connections.teradata import TeradataHandler
from pythonwf.logging.logging import call_logger, CustomLogger
from pythonwf.eligibility.eligibility import Eligible
import pandas as pd
from collections import OrderedDict
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime


class Waterfall:
    """
    A class to generate and analyze waterfall reports for campaign eligibility checks.

    Attributes:
        conditions (pd.DataFrame): Conditions for eligibility checks.
        offer_code (str): The offer code.
        campaign_planner (str): The campaign planner.
        lead (str): The lead person.
        waterfall_location (str): The location to save the waterfall report.
        _sqlconstructor (SQLConstructor): An instance of SQLConstructor to build SQL queries.
        _teradata_connection (Optional[TeradataHandler]): The Teradata connection handler.
        _column_names (OrderedDict): Column names for the waterfall report.
        _query_results (Dict[str, List[pd.DataFrame]]): Query results for each identifier.
        _compiled_dataframes (OrderedDict): Compiled dataframes for each identifier.
        _starting_population (Optional[int]): The starting population for the waterfall analysis.
        _combined_df (Optional[pd.DataFrame]): Combined dataframe for the waterfall report.
    """

    def __init__(
            self,
            conditions: Dict[str, Dict[str, Any]],
            offer_code: str,
            campaign_planner: str,
            lead: str,
            waterfall_location: str,
            sql_constructor: SQLConstructor,
            logger: CustomLogger,
            teradata_connection: Optional[TeradataHandler] = None
    ) -> None:
        """
        Initializes the Waterfall class with the provided parameters.

        Args:
            conditions (Dict[str, Dict[str, Any]]): Conditions for eligibility checks.
            offer_code (str): The offer code.
            campaign_planner (str): The campaign planner.
            lead (str): The lead person.
            waterfall_location (str): The location to save the waterfall report.
            sql_constructor (SQLConstructor): An instance of SQLConstructor to build SQL queries.
            teradata_connection (Optional[TeradataHandler]): The Teradata connection handler.
        """
        self.logger = logger
        self.current_date = datetime.now().strftime("%Y%m%d")

        self._sqlconstructor = sql_constructor
        self._teradata_connection = teradata_connection

        self.offer_code = offer_code
        self.campaign_planner = campaign_planner
        self.lead = lead
        self.waterfall_location = waterfall_location

        self.conditions = conditions

        # prep column names
        self._column_names = OrderedDict({
            'unique_drops': '{identifier} drop if only this drop',
            'increm_drops': '{identifier} drop increm',
            'cumul_drops': '{identifier} drop cumul',
            'regain': '{identifier} regain if no scrub',
            'remaining': '{identifier} remaining'
        })

        # initializing properties
        self._query_results = dict()
        self._compiled_dataframes = OrderedDict()
        self._starting_population = None
        self._combined_df = None

    @classmethod
    def from_eligible(cls, eligibility: Eligible, waterfall_location: str) -> 'Waterfall':
        conditions = cls._prepare_conditions(eligibility.conditions)
        offer_code = eligibility.offer_code
        campaign_planner = eligibility.campaign_planner
        lead = eligibility.lead
        sql_constructor = eligibility._sqlconstructor
        logger = eligibility.logger
        teradata_connection = eligibility._teradata_connection

        return cls(conditions, offer_code, campaign_planner, lead, waterfall_location, sql_constructor, logger, teradata_connection)

    @classmethod
    def _prepare_conditions(cls, conditions: dict) -> dict:
        """
        Prepares the conditions by transforming them into a dictionary.

        Args:
            conditions (Dict[str, Dict[str, Any]]): Conditions for eligibility checks.

        Returns:
            dict: The prepared conditions as a dictionary.
        """
        result_dict = {}

        for channel, templates in conditions.items():
            for template, checks in templates.items():
                for check in checks:
                    column_name = check.get('column_name', None)
                    description = check.get('description', None)
                    sql = check.get('sql', None)

                    if column_name is not None:
                        modified_description = f'[{template}] {description}' if description else None
                        result_dict[column_name] = {
                            'description': modified_description,
                            'sql': sql
                        }

        return result_dict

    @property
    def conditions(self) -> dict:
        """Getter for conditions."""
        return self._conditions

    @conditions.setter
    def conditions(self, value: Dict[str, Dict[str, Any]]) -> None:
        """Setter for conditions."""
        self._conditions = value

    def _save_results(self, identifier: str, data: pd.DataFrame) -> None:
        """
        Saves the results of a query to the _query_results dictionary.

        Args:
            identifier (str): The identifier for the query results.
            data (pd.DataFrame): The data to save.
        """
        if self._query_results.get(identifier) is None:
            self._query_results[identifier] = []
        self._query_results[identifier].append(data)

    @call_logger()
    def _calculate_regain(self) -> None:
        """
        Calculates the regain SQL and saves the results to the _query_results dictionary.
        """
        queries = self._sqlconstructor.waterfall.generate_regain_sql()

        # save queries
        self.logger.info(f'{self.__class__}._calculate_regain {queries=}')

        for identifier, query in queries.items():
            if self._teradata_connection is not None:
                df = self._teradata_connection.to_pandas(query)
                df['Index'] = self._column_names.get('regain').format(identifier=identifier)
                df = df.set_index('Index')
                self._save_results(identifier, df)

    @call_logger()
    def _calculate_incremental_drops(self) -> None:
        """
        Calculates the incremental drops SQL and saves the results to the _query_results dictionary.
        """
        queries = self._sqlconstructor.waterfall.generate_incremental_drops_sql()

        # save queries
        self.logger.info(f'{self.__class__}._calculate_incremental_drops {queries=}')

        for identifier, query in queries.items():
            if self._teradata_connection is not None:
                df = self._teradata_connection.to_pandas(query)
                df['Index'] = self._column_names.get('increm_drops').format(identifier=identifier)
                df = df.set_index('Index')
                self._save_results(identifier, df)

    @call_logger()
    def _calculate_unique_drops(self) -> None:
        """
        Calculates the unique drops SQL and saves the results to the _query_results dictionary.
        """
        queries = self._sqlconstructor.waterfall.generate_unique_drops_sql()

        # save queries
        self.logger.info(f'{self.__class__}._calculate_unique_drops {queries=}')

        for identifier, query in queries.items():
            if self._teradata_connection is not None:
                df = self._teradata_connection.to_pandas(query)
                df['Index'] = self._column_names.get('unique_drops').format(identifier=identifier)
                df = df.set_index('Index')
                self._save_results(identifier, df)

    @call_logger()
    def _calculate_remaining(self) -> None:
        """
        Calculates the remaining SQL and saves the results to the _query_results dictionary.
        """
        queries = self._sqlconstructor.waterfall.generate_remaining_sql()

        # save queries
        self.logger.info(f'{self.__class__}._calculate_remaining {queries=}')

        for identifier, query in queries.items():
            if self._teradata_connection is not None:
                df = self._teradata_connection.to_pandas(query)
                df['Index'] = self._column_names.get('remaining').format(identifier=identifier)
                df = df.set_index('Index')
                self._save_results(identifier, df)

    def _step1_create_base_tables(self):
        queries = self._sqlconstructor.waterfall.generate_unique_identifier_details_sql()

        self.logger.info(f'{self.__class__}.step1_create_base_tables {queries=}')

        for identifier, details in queries.items():
            query = details.get('sql')
            table_name = details.get('table_name')

            if self._teradata_connection is not None:
                self._teradata_connection.execute_query(query)
                self._teradata_connection.tracking.track_table(table_name)

    @call_logger()
    def _step2_analyze_eligibility(self) -> None:
        """
        Analyzes the eligibility by calculating unique drops, incremental drops, regain, and remaining records.
        The order of calculations is important for correct results.
        """
        self._calculate_unique_drops()
        self._calculate_incremental_drops()
        self._calculate_regain()
        self._calculate_remaining()

    @call_logger
    def _step3_create_dataframes(self) -> None:
        """
        Creates dataframes from the query results and compiles them into a dictionary.
        """
        for identifier, dfs in self._query_results.items():
            df = pd.concat(dfs, axis=1).transpose()

            unique_drop = self._column_names.get('unique_drops').format(identifier=identifier)
            increm_drop = self._column_names.get('increm_drops').format(identifier=identifier)
            cumul_drop = self._column_names.get('cumul_drops').format(identifier=identifier)
            regain = self._column_names.get('regain').format(identifier=identifier)
            remaining = self._column_names.get('remaining').format(identifier=identifier)

            self._starting_population = df.loc[0, increm_drop] + df.loc[0, remaining]
            df[cumul_drop] = self._starting_population - df[remaining]

            df = df[[unique_drop, increm_drop, cumul_drop, regain, remaining]]
            self._compiled_dataframes[identifier] = df

    @call_logger
    from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font

def _step4_create_excel(self):
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Waterfall'

    # Add header information
    header = f'[{self.offer_code}] [CP: {self.campaign_planner}] [LEAD: {self.lead}] [DATE: {self.current_date}]'
    ws['A1'] = header
    ws['A1'].font = Font(size=18)

    # Add values to specific cells
    ws['A2'] = 'Checks'
    ws['B2'] = 'Criteria'
    ws['C2'] = 'Description'
    ws['C3'] = 'Starting Population'

    # Write the first dataframe values starting from cell A4
    for r_idx, row in enumerate(dataframe_to_rows(pd.DataFrame(self.conditions), index=False, header=False), start=4):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Write other dataframes starting from appropriate columns and add headers
    start_col = 5
    for key, df in self._query_results.items():
        for col_num, value in enumerate(df.columns, start=start_col):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=4):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)
        start_col += len(df.columns) + 1

    # Insert blank row when channel name changes and rename values in column A
    df_combined = pd.concat([self.conditions] + list(self._query_results.values()), ignore_index=True)
    df_combined.reset_index(drop=True, inplace=True)

    channel_name_prev = ''
    for idx, row in df_combined.iterrows():
        if pd.isna(row['column_name']):
            continue
        channel_name, template, check_number = row['column_name'].split('_', 2)
        if channel_name != channel_name_prev and idx != 0:
            ws.insert_rows(idx + 5)
        ws.cell(row=idx + 5, column=1, value=check_number)
        channel_name_prev = channel_name

    # Loop through column A and insert a blank row when channel name changes, then leave only the checknumber
    channel_name_prev = ''
    for row in range(4, len(df_combined) + 5):
        cell = ws.cell(row=row, column=1)
        if cell.value is None:
            continue
        parts = cell.value.split('_')
        channel_name = parts[0]
        check_number = parts[-1]
        if channel_name != channel_name_prev and row != 4:
            ws.insert_rows(row)
        cell.value = check_number
        channel_name_prev = channel_name

    # Formatting
    tan_fill = PatternFill(start_color='D2B48C', end_color='D2B48C', fill_type='solid')
    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    header_fill = PatternFill(start_color='01204E', end_color='01204E', fill_type='solid')
    col_a_fill = PatternFill(start_color='F6DCAC', end_color='F6DCAC', fill_type='solid')
    blank_fill = PatternFill(start_color='028391', end_color='028391', fill_type='solid')

    # Apply header fill color to row 2
    for col in range(1, start_col):
        ws.cell(row=2, column=col).fill = header_fill

    # Apply color to column A starting from row 3
    for row in range(3, len(df_combined) + 5):
        ws.cell(row=row, column=1).fill = col_a_fill

    # Apply color to blank rows and columns
    for row in range(4, len(df_combined) + 5):
        if ws.cell(row=row, column=2).value is None:
            for col in range(1, start_col):
                ws.cell(row=row, column=col).fill = blank_fill

    for col in range(1, start_col):
        if ws.cell(row=4, column=col).value is None:
            for row in range(3, len(df_combined) + 5):
                ws.cell(row=row, column=col).fill = blank_fill

    # Save the Excel file
    wb.save(f'{self.waterfall_location}/{self.offer_code}_Waterfall_{self.current_date}.xlsx')


    def generate_waterfall(self):
        self._step1_create_base_tables()
        self._step2_analyze_eligibility()
        self._step3_create_dataframes()
        self._step4_create_excel()
